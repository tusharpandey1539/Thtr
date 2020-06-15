import sys
import datetime
from openpyxl import *
from pip._vendor.distlib.compat import raw_input
from PIL import Image

loc = 'thtr.xlsx'
locMov = 'mov.xlsx'
locSeat = 'seating.xlsx'
global mov_title

print("\n****************************************ThTR********"
      "********************************\n")


def reg_user():
    wb = load_workbook(loc)
    sheet = wb.active
    current_row = sheet.max_row
    print("\n--------------------Register--------------------\n")

    sheet.cell(row=current_row + 1,
               column=1).value = input("          Enter the UserName: ")

    e_mail_id = input("          Enter the Email Id: ")
    if "@" in e_mail_id and "." in e_mail_id:
        sheet.cell(row=current_row + 1, column=2).value = e_mail_id
    else:
        print("     ERROR: WRONG INPUT. PLEASE TRY AGAIN.....")
        reg_user()

    sheet.cell(row=current_row + 1,
               column=3).value = input("          Enter the PassWord: ")

    p_number = input("          Enter the PhoneNo.: (+91) ")
    if len(p_number) == 10:
        sheet.cell(row=current_row + 1,
                   column=4).value = p_number
    else:
        print("     ERROR: WRONG INPUT. PLEASE TRY AGAIN.....")
        reg_user()

    wb.save(loc)


def login_user():
    wb = load_workbook(loc)
    sheet = wb.active
    current_row = sheet.max_row
    # current_column = sheet.max_column
    print("\n--------------------Login--------------------\n")
    Username = input("          Enter the UserName: ")
    Password = input("          Enter the PassWord: ")
    for i in range(1, current_row + 1):
        if sheet.cell(row=i, column=1).value == Username:
            if sheet.cell(row=i, column=3).value == Password:
                return True
    wb.save(loc)
    return False


def login_menu():
    print("\n--------------------Main Menu--------------------\n")
    choice = input('\n'
                   '          1: Register / Sign Up\n'
                   '          2: Login / Sign In\n'
                   '          Q: Quit/Log Out\n'
                   '          Please enter your choice: ')

    if choice == "1":
        reg_user()
        login_menu()
    elif choice == "2":
        if login_user():
            main_menu()
        else:
            login_menu()
    elif choice == "Q":
        sys.exit()
    else:
        print("     ERROR: WRONG INPUT. PLEASE TRY AGAIN.....")
        login_menu()


def location():
    print("\n--------------------Location Finder--------------------\n")
    print("          Popular locations")
    print("          1: Bangalore\n"
          "          2: Delhi\n"
          "          3: Jaipur\n"
          "          4: Chennai\n"
          "          5: Mumbai")

    city: str = input("          Please enter your choice\n"
                      "          If you can't find your desired location "
                      "then please go ahead and\n"
                      "          Enter the name of the location:  ")
    return city


def display():
    print("\n--------------------Movie Directory--------------------\n")
    wbm = load_workbook(locMov)
    sheets = wbm.active
    current_rowz = sheets.max_row
    mov_choice = input("          1: View Trending Movies:\n"
                       "          2: View all currently running movies:\n"
                       "          3: Search your desired Movie:\n"
                       "          4: Filter movie by choice:\n"
                       "          Enter Your Choice: ")
    if mov_choice == "1":
        for i in range(2, 6):
            print("          ", sheets.cell(row=i, column=1).value)
    elif mov_choice == "2":
        for i in range(2, current_rowz + 1):
            print("          ", sheets.cell(row=i, column=1).value)
    elif mov_choice == "3":
        pass
    elif mov_choice == "4":
        wcat = input("          Filter by:\n"
                     "            1: Genre\n"
                     "            2: Language\n"
                     "          Enter your choice: ")
        if wcat == "1":
            genre_choice = input("          Available: 1: Action\n"
                                 "                     2: Comedy\n"
                                 "                     3: Drama\n"
                                 "                     4: Horror\n"
                                 "                     5: Romance\n"
                                 "                     6: Sci-fi\n"
                                 "          Enter your choice: ")
            for i in range(2, current_rowz + 1):
                if sheets.cell(row=i, column=4).value == genre_choice:
                    print(sheets.cell(row=i, column=1).value)

        if wcat == "2":
            lang_choice = input("          Available: 1: English\n"
                                "                     2: Hindi\n"
                                "                     3: Telugu\n"
                                "          Enter your choice: ")
            for i in range(2, current_rowz + 1):
                if sheets.cell(row=i, column=5).value == lang_choice:
                    print(sheets.cell(row=i, column=1).value)

        else:
            print("     ERROR: WRONG INPUT. PLEASE TRY AGAIN.....")
            display()
    else:
        print("     ERROR: WRONG INPUT. PLEASE TRY AGAIN.....")
        display()

    wbm.save(locMov)


def details():
    print("\n--------------------Movie Details--------------------\n")
    global mov_title
    mov_name = input("          Enter the movie name: ")
    wbm = load_workbook(locMov)
    sheets = wbm.active
    current_row = sheets.max_row
    flag = 0
    for i in range(1, current_row + 1):
        if sheets.cell(row=i, column=1).value == mov_name:
            print("          Movie Title: ", sheets.cell(row=i, column=1).value)
            print("          Movie Description: ", sheets.cell(row=i, column=2).value)
            print("          Movie Rating: ", sheets.cell(row=i, column=3).value)
            print("          Movie Genre: ", sheets.cell(row=i, column=4).value)
            print("          Movie Language: ", sheets.cell(row=i, column=5).value)
            print("          Movie Cast: ", sheets.cell(row=i, column=6).value)
            if input("Show Movie Poster (y/n): ") == "y":
                img = Image.open(sheets.cell(row=i, column=7).value)
                img.show()
            else:
                pass
            print("          Movie Trailer Link: ", sheets.cell(row=i, column=8).value)
            flag = 1
            mov_title = sheets.cell(row=i, column=1).value
    wbm.save(locMov)
    if flag == 0:
        print("     ERROR: NO SUCH MOVIE EXISTS")
        print("     PLEASE TRY AGAIN.....")
        details()
    return mov_title


def theatre():
    print("\n--------------------Theatre Selector--------------------\n")
    theatre_name = input("          Available theatres in your area:\n"
                         "          1: INOX\n"
                         "          2: PVR\n"
                         "          3: GOLD\n"
                         "          4: FUN CINEMAS\n"
                         "          Enter your Choice (Theatre name):  ")
    return theatre_name


def time():
    print("\n--------------------Show Selector--------------------\n")
    y, m, dd = map(int, raw_input("          Enter the date in the"
                                  " format y,m,d: ").split(","))
    current_date = datetime.datetime.now()
    typed_date = datetime.datetime(y, m, dd)
    if typed_date >= current_date:
        print("          Available time slots on selected date:\n"
              "          1: 09:00:00"
              "          2: 12:30:00"
              "          3: 16:00:00"
              "          4: 19:30:00"
              "          5: 23:00:00 ")
        h, mi, s = map(int, raw_input("          Enter the date in the "
                                      "format h:m:s: ").split(":"))
        date_time = datetime.datetime(y, m, dd, h, mi, s)
        if date_time >= current_date:
            return date_time
    else:
        print("     ERROR: WRONG INPUT. PLEASE TRY AGAIN....")
        time()


def seating():
    print("\n--------------------Seat Selector--------------------\n")
    wbs = load_workbook(locSeat)
    sheet_seating = wbs.active
    current_row = sheet_seating.max_row
    current_col = sheet_seating.max_column
    print('\n'.join(['\t'.join([str(sheet_seating.cell(row=i, column=j).value)
                                for j in range(1, current_col + 1)])
                     for i in range(1, current_row + 1)]))
    print("--------------Display this side------------------ ")
    print("         Note: Booked seats are displayed with a star.")
    booked_seats = list(map(str, input("          Enter seats you"
                                       " want to book: ").split()))
    return booked_seats


def promo_codes():
    print("\n--------------------Promo Codes--------------------\n")
    print("          Available Promo-codes for today:\n"
          "          1:   R4S5s7i6  : free vouchers for food store\n"
          "          2:   seXH8S11    : 50 Rs off\n"
          "          3:   Qf3olL54    : 10% discount")
    choice_promo = input("          Enter your choice: ")
    if choice_promo == "1":
        return 55
    elif choice_promo == "2":
        return 48
    elif choice_promo == "3":
        return 35
    else:
        return 0


def payment_amount(choice_promo_code, seating_array):
    price = (len(seating_array)) * 200
    if choice_promo_code == 55:
        return price
    elif choice_promo_code == 48:
        price = price - 50
        return price
    elif choice_promo_code == 35:
        price = price - (price / 10)
        return price
    else:
        return price


def payment_options():
    print("\n--------------------Payment Page--------------------\n")
    pay_opt = input("          Enter the type of payment:\n"
                    "          1: Credit/Debit Card\n"
                    "          2: Upi\n"
                    "          Your choice: ")
    if pay_opt == "1":
        print("          You chose to pay by card:\n")
        input("          Enter your 16 digit card number: ")
        input("          Enter your 3 digit cvv: ")
        input("          Enter your month "
              "and year of expiry (in format MM/YY): \n")
        print("          An otp has been sent"
              " to your registered mobile number.\n")
        input('          Enter the otp sent your registered mobile no.: ')
        print("          ....Payment processing.......")

    elif pay_opt == "2":
        print("          You chose to pay by UPI:\n"
              "          wait.....\n"
              "          A payment request has been"
              " sent to your registered mobile no.\n"
              "          accept the request to proceed with payment\n"
              "          ....Payment processing......")
    else:
        print("     ERROR: WRONG INPUT. PLEASE TRY AGAIN")
        payment_options()


def booking(final_location, final_movie_name, final_theatre_name,
            final_time_date, final_array_seats, final_payment_amount):
    print("\n--------------------Booking Menu--------------------\n")
    print("         Location:  ", final_location)
    print("         Movie:  ", final_movie_name)
    print("         Theatre:  ", final_theatre_name)
    print("         Date and Time of Show:  ", final_time_date)
    print("         Seats Booked:  ", final_array_seats)
    print("         Payment amount:  ", final_payment_amount)
    choice = input("          Continue to payment (y/n): ")
    if choice == "y" or choice == "Y":
        payment_options()
        print("\n--------------------Invoice--------------------\n")
        print("         Booking confirm")
        print("         Location:  ", final_location)
        print("         Movie:  ", final_movie_name)
        print("         Theatre:  ", final_theatre_name)
        print("         Date and Time of Show:  ", final_time_date)
        print("         Seats Booked:  ", final_array_seats)
        print("         Paid amount:  ", final_payment_amount)
    else:
        main_menu()


def logout():
    login_menu()


def main_menu():
    location_runtime = location()
    print("\n--------------------Main Menu--------------------\n")
    choice = input("          1: Display/ Search Movies/ Book tickets\n"
                   "          2: Logout\n"
                   "          Enter your Choice: ")
    if choice == "1":
        display()
        movie_name_runtime = details()
        theatre_name_runtime = theatre()
        time_runtime = time()
        seats_runtime = seating()
        promo_runtime = promo_codes()
        price_runtime = payment_amount(promo_runtime, seats_runtime)
        booking(location_runtime, movie_name_runtime, theatre_name_runtime,
                time_runtime, seats_runtime, price_runtime)
        exit_now = input("         Exit Now??(y/n): ")
        if exit_now == "y" or exit_now == "Y":
            sys.exit()
        else:
            main_menu()
    elif choice == "2":
        logout()
    else:
        print("     ERROR: WRONG INPUT. PLEASE TRY AGAIN.....")
        main_menu()


login_menu()

